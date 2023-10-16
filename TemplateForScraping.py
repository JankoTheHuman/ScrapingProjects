#THIS FILE SERVES ONLY AS A TEMPLATE FOR FUTURE SCRAPING PROJECTS FOR AN EASIER START

import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
import time
from selenium.common.exceptions import StaleElementReferenceException, NoSuchElementException

RepeatProgram = True
j=0
k=0

while RepeatProgram:
    try:
        chrome_options = webdriver.ChromeOptions()
        driver = webdriver.Chrome(options=chrome_options)
        sheet_name = 'Sheet1'
        chrome_options.add_argument("--start-maximized")
        ##############driver.get("LINK")

        ############## excel_file = r'C:\Users\Asus\Desktop\New folder\New.xlsx'

        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook[sheet_name]
        next_row = sheet.max_row + 1



        for i in range(100000000):

            newRow = {'Company Name': comp, 'State': linksToScrape[i][-2:]}
            newRow[0] = 1

            # prints data into Excel
            for col, value in enumerate(newRow.values(), start=1):
                sheet.cell(row=next_row, column=col, value=value)
            next_row += 1

            # data excel
            workbook.save(excel_file)



    except NoSuchElementException:
        workbook.save(excel_file)
        driver.quit()
        print(f"No Such Elemnt Exception raised, outsidemost try/catch. This has happened {j} times.")

    except StaleElementReferenceException:
        workbook.save(excel_file)
        driver.quit()
        print(f"Stale Element Reference Exception raised, outsidemost try/catch. This has happened {k} times.")