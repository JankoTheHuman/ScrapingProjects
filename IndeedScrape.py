#THIS FILE SERVES ONLY AS A TEMPLATE FOR FUTURE SCRAPING PROJECTS FOR AN EASIER START
from random import random

import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
import time
from selenium.common.exceptions import StaleElementReferenceException, NoSuchElementException

RepeatProgram = True
j=0
k=0

SEARCH_CRITERIA= "BDR"
PAGE_START= 10
pagecheck=0

while RepeatProgram:
    try:
        chrome_options = webdriver.ChromeOptions()
        driver = webdriver.Chrome(options=chrome_options)
        sheet_name = 'Sheet1'
        chrome_options.add_argument("--start-maximized")

        excel_file = r'C:\Users\janko\OneDrive\Desktop\Indeed SDR search.xlsx'

        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook[sheet_name]
        next_row = sheet.max_row + 1




        for i in range(100000000):

            driver.get(f"https://www.indeed.com/jobs?q={SEARCH_CRITERIA}&start={PAGE_START}")
            time.sleep(random()+5+random())

            mozaic = driver.find_element(By.ID,"mosaic-jobResults")
            jobs= mozaic.find_elements(By.CSS_SELECTOR,"div.cardOutline.tapItem.dd-privacy-allow.result")


            for job in jobs:

                title= job.find_element(By.TAG_NAME,'span').text
                comp= job.find_element(By.CSS_SELECTOR,'span.css-1x7z1ps.eu4oa1w0').text
                location= job.find_element(By.CSS_SELECTOR,'div.css-t4u72d.eu4oa1w0').text
                date= job.find_element(By.CSS_SELECTOR,'span.date').text



                newRow = {'Vacancy title':title, 'Company Name': comp, 'Location':location, 'date':date}
                print(newRow)

                # prints data into Excel
                for col, value in enumerate(newRow.values(), start=1):
                    sheet.cell(row=next_row, column=col, value=value)
                next_row += 1

                # save data excel
                workbook.save(excel_file)

            pagecheck+=15
            PAGE_START+=10


            #check if you reached the end
            jobs_found = driver.find_element(By.CSS_SELECTOR,".jobsearch-JobCountAndSortPane-jobCount.css-1af0d6o.eu4oa1w0")
            jobs_found = jobs_found.find_element(By.TAG_NAME,"span").text
            jobs_found = jobs_found.split()[0]

            print(f"{pagecheck}-------{jobs_found}\n")


            if pagecheck > int(jobs_found):
                exit()


    except NoSuchElementException:
        workbook.save(excel_file)
        driver.get(f"https://www.indeed.com/jobs?q={SEARCH_CRITERIA}&start={PAGE_START}")
        j+=1
        print(f"No Such Elemnt Exception raised, outsidemost try/catch. This has happened {j} times.")

    except StaleElementReferenceException:
        workbook.save(excel_file)
        driver.get(f"https://www.indeed.com/jobs?q={SEARCH_CRITERIA}&start={PAGE_START}")
        k+=1
        print(f"Stale Element Reference Exception raised, outsidemost try/catch. This has happened {k} times.")