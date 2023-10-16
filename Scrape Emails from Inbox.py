
#This project scrapes email from inbox, only changable values are EMAIL, password and search from which
# you want to scrape emails

import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.select import Select
import time
import random
from selenium.common.exceptions import StaleElementReferenceException, NoSuchElementException

#---------------------------------------------------------------------------
#CHANGE THIS PART ONLY
#NOTE: Messages have to be marked as unread in the email
EMAIL = "emaill@email.com"
PASSWORD = "password"
SEARCH = "missed you at sfr"

SEARCHLINK = SEARCH.replace(" ","+")
SEARCH = "in%3Ainbox+" + SEARCH
#---------------------------------------------------------------------------

j=0
l=0
m=1
excel_file = r'C:\Users\Asus\Desktop\New folder\EmailScrape.xlsx'
sheet_name = 'Sheet1'

workbook = openpyxl.load_workbook(excel_file)
sheet = workbook[sheet_name]

next_row = sheet.max_row + 1
openProgramAgain = True

while openProgramAgain:
    try:
        openProgramAgain = True
        chrome_options = webdriver.ChromeOptions()
        chrome_options.add_argument("--start-maximized")
        driver = webdriver.Chrome(options=chrome_options)
        driver.get("https://mail.google.com/mail/u/0/#inbox")
        time.sleep(2)

        print(f'Times opened program{m}--------------------')
        m+=1
        #email input
        time.sleep(2)
        driver.find_element(By.ID, "identifierId").send_keys(EMAIL)

        # first next button
        time.sleep(2)
        driver.find_element(By.CSS_SELECTOR, "button.VfPpkd-LgbsSe.VfPpkd-LgbsSe-OWXEXe-k8QpJ.VfPpkd-LgbsSe-OWXEXe-dgl2Hf.nCP5yc.AjY5Oe.DuMIQc.LQeN7.qIypjc.TrZEUc.lw1w4b").click()

        #password input
        time.sleep(2)
        driver.find_element(By.CSS_SELECTOR,".whsOnd.zHQkBf").send_keys(PASSWORD)

        # second next button
        time.sleep(3)
        driver.find_element(By.CSS_SELECTOR, "button.VfPpkd-LgbsSe.VfPpkd-LgbsSe-OWXEXe-k8QpJ.VfPpkd-LgbsSe-OWXEXe-dgl2Hf.nCP5yc.AjY5Oe.DuMIQc.LQeN7.qIypjc.TrZEUc.lw1w4b").click()

        time.sleep(5)

        try:
            driver.find_element(By.CSS_SELECTOR,'input.whsOnd.zHQkBf')
            print("Google Authenticator CODE Required")

            for i in range(30):
                print(i)
                time.sleep(1)
        except NoSuchElementException:
            print("GAC not required")


        # HERE SCRAPE EMAIL's and switch pages
        try:
            for i in range(1,9999):
                driver.get(f"https://mail.google.com/mail/u/0/#search/{SEARCH}/p{i}")
                time.sleep(10)

                div = driver.find_elements(By.CSS_SELECTOR,".F.cf.zt")
                tableRows = div[1].find_elements(By.TAG_NAME,'tr')

                for row in tableRows:

                    message = row.find_elements(By.CSS_SELECTOR, "span.zF")
                    date = row.find_element(By.CSS_SELECTOR, 'span.bq3').text
                    name = message[0].get_attribute('name')
                    email = message[0].get_attribute('email')

                    if len(message)<3:
                        newRow = {'name': name , 'email': email, 'date': date}

                        for col, value in enumerate(newRow.values(), start=1):
                            sheet.cell(row=next_row, column=col, value=value)
                        next_row += 1
                        workbook.save(excel_file)

                    else:
                        for s in range(0,int(len(message)/2)):
                            name = message[s].get_attribute('name')
                            email = message[s].get_attribute('email')
                            newRow = {'name': name, 'email': email, 'date': date}

                            for col, value in enumerate(newRow.values(), start=1):
                                sheet.cell(row=next_row, column=col, value=value)
                            next_row += 1
                            workbook.save(excel_file)



        except NoSuchElementException:
            openProgramAgain = False
            print('No such Element exception -> probably end of the list/search')


        driver.quit()

    except StaleElementReferenceException:
        driver.quit()
        j += 1
        print("RAISED StaleElementReferenceException-------------------------------------------" + str(j) + "Times")

    except NoSuchElementException:
        driver.quit()
        l += 1
        print("RAISED NoSuchElementException  -------------------------------------------" + str(l) + "Times")
