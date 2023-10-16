
#THIS project was served to scrape Companies and their info from nplaconference.com
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
import time
from selenium.common.exceptions import StaleElementReferenceException, NoSuchElementException

repeatProgram = True
j=0
k=0
pageReached=1

while repeatProgram:
    try:
        chrome_options = webdriver.ChromeOptions()
        driver = webdriver.Chrome(options=chrome_options)
        sheet_name = 'Sheet1'
        chrome_options.add_argument("--start-maximized")
        excel_file = r'C:\Users\Asus\Desktop\New folder\NPLA Directory.xlsx'

        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook[sheet_name]
        next_row = sheet.max_row + 1



        for i in range(1,18):

            driver.get(f"https://nplaconference.com/private-lending-directory/page{i}/?listingcateg=-1&listingregion=0&frompaged=true")
            for t in range(5):
                print(t)
                time.sleep(1)
            print(f'Doing page {i}')

            listingSpace = driver.find_element(By.ID, "thelistingsplace")
            topSide = listingSpace.find_elements(By.CSS_SELECTOR, "div.topside")
            bottomSide = listingSpace.find_elements(By.CSS_SELECTOR, "div.bottomside")

            for n in range(len(topSide)):

                a = bottomSide[n].find_elements(By.TAG_NAME,'a')

                li= topSide[n].find_elements(By.TAG_NAME,'li')
                email=''
                company = topSide[n].find_element(By.TAG_NAME,"h3").text
                location =li[0].text[1:]
                try:
                    email = a[1].get_attribute('href')[7:]
                except:
                    ''
                website = a[0].get_attribute('href')
                phone = li[1].text[1:]

                newRow = {'Company Name': company, 'Location': location, 'Company Email': email, 'website':website, 'phone': phone}
                print(newRow)

                # prints data into Excel
                for col, value in enumerate(newRow.values(), start=1):
                    sheet.cell(row=next_row, column=col, value=value)
                next_row += 1

                # data excel
                workbook.save(excel_file)

            if i == 17:
                repeatProgram = False

        driver.quit()

    except  NoSuchElementException:
        workbook.save(excel_file)
        driver.quit()
        print(f"No Such Elemnt Exception raised, outsidemost try/catch. This has happened {j} times.")

    except StaleElementReferenceException:
        workbook.save(excel_file)
        driver.quit()
        print(f"Stale Element Reference Exception raised, outsidemost try/catch. This has happened {k} times.")




















